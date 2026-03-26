import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook(r'd:\remina\remina_roi_sample.xlsx', data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*80}")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}")

    # Tab color
    if ws.sheet_properties and ws.sheet_properties.tabColor:
        print(f"Tab Color: {ws.sheet_properties.tabColor.rgb}")

    # Merged cells
    print(f"\n--- MERGED CELLS ---")
    if ws.merged_cells.ranges:
        for merge in ws.merged_cells.ranges:
            print(f"  {merge}")
    else:
        print("  (none)")

    # Column widths
    print(f"\n--- COLUMN WIDTHS ---")
    for col_letter, col_dim in sorted(ws.column_dimensions.items()):
        print(f"  Col {col_letter}: width={col_dim.width}, hidden={col_dim.hidden}, bestFit={col_dim.bestFit}")

    # Row heights
    print(f"\n--- ROW HEIGHTS ---")
    for row_num, row_dim in sorted(ws.row_dimensions.items()):
        print(f"  Row {row_num}: height={row_dim.height}, hidden={row_dim.hidden}")

    # Sheet-level print/page settings
    print(f"\n--- PAGE SETUP ---")
    print(f"  page_margins: left={ws.page_margins.left}, right={ws.page_margins.right}, top={ws.page_margins.top}, bottom={ws.page_margins.bottom}")
    print(f"  freeze_panes: {ws.freeze_panes}")
    print(f"  sheet_view zoom: {ws.sheet_view.zoomScale if ws.sheet_view else 'N/A'}")

    # Cell details - include ALL cells that have a value OR non-default style
    print(f"\n--- CELL DETAILS ---")
    for row in ws.iter_rows():
        for cell in row:
            has_value = cell.value is not None

            # Check fill
            fill = cell.fill
            fill_type = fill.fill_type
            has_fill = fill_type not in (None, 'none')

            # Check font customization
            font = cell.font
            has_font_style = (font.bold or font.italic or font.underline or
                              (font.size and font.size != 11) or
                              (font.name and font.name not in ('Calibri', None)) or
                              font.strike)

            # Check borders
            border = cell.border
            def has_border_side(side):
                return side and side.border_style is not None
            has_border = any(has_border_side(s) for s in [border.top, border.bottom, border.left, border.right])

            # Check alignment
            align = cell.alignment
            has_align = any([align.horizontal, align.vertical, align.wrap_text, align.indent])

            # Skip truly empty unstyled cells
            if not any([has_value, has_fill, has_font_style, has_border, has_align]):
                continue

            addr = cell.coordinate
            val = cell.value

            # Fill colors
            def color_str(color_obj):
                if color_obj is None:
                    return "N/A"
                try:
                    if color_obj.type == 'rgb':
                        return color_obj.rgb
                    elif color_obj.type == 'theme':
                        return f"theme:{color_obj.theme}(tint:{color_obj.tint})"
                    elif color_obj.type == 'indexed':
                        return f"indexed:{color_obj.indexed}"
                    else:
                        return f"type:{color_obj.type}"
                except Exception as e:
                    return f"err:{e}"

            fg_color = color_str(fill.fgColor)
            bg_color = color_str(fill.bgColor)

            # Font color
            font_color = color_str(font.color)

            # Borders
            def border_info(side):
                if side and side.border_style:
                    return f"{side.border_style}/{color_str(side.color)}"
                return "none"

            b_top    = border_info(border.top)
            b_bottom = border_info(border.bottom)
            b_left   = border_info(border.left)
            b_right  = border_info(border.right)
            b_diag   = border_info(border.diagonal)

            num_fmt = cell.number_format

            # Data type
            dtype = cell.data_type

            print(f"\n  [{addr}] value={repr(val)}  data_type={dtype}")
            print(f"    FILL:   fill_type={fill_type!r}, fgColor={fg_color}, bgColor={bg_color}")
            print(f"    FONT:   bold={font.bold}, italic={font.italic}, size={font.size}, name={font.name!r}, color={font_color}, underline={font.underline}, strike={font.strike}")
            print(f"    ALIGN:  horiz={align.horizontal!r}, vert={align.vertical!r}, wrap={align.wrap_text}, indent={align.indent}, shrink={align.shrink_to_fit}, rotate={align.text_rotation}")
            print(f"    FORMAT: {repr(num_fmt)}")
            print(f"    BORDER: top={b_top}, bottom={b_bottom}, left={b_left}, right={b_right}, diag={b_diag}")

print("\n\n=== EXTRACTION COMPLETE ===")
